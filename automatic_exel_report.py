import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import sys
import time

month = input('Enter Month: ')
app_path = os.path.dirname(sys.executable)
app_path = os.path.dirname(os.path.abspath(__file__))
file_location = os.path.join(app_path,'Product-Sales-Region.xlsx')
sample_file_location = os.path.join(app_path,f'sample_chart_{month}.xlsx')
print(app_path)


def create_sample_data(file_name):
    try:
        df = pd.read_excel(file_name)
        sample_table = df.pivot_table(index='Product',columns='CustomerType',values='TotalPrice',aggfunc='sum')
        sample_table.to_excel(f'sample_chart_{month}.xlsx','Report',startrow=4)
        print("Sample Data created!")

    except Exception:
        print("File not found!")
    # print(df[['Product','CustomerType','TotalPrice']])

    


def sumofEach():
    wb = load_workbook(sample_file_location)

    sheet = wb['Report']

    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row= wb.active.max_row

    for i in range(min_row+1,max_row+1):
        # print(i)
        # print(get_column_letter(i))
        sheet[f'{get_column_letter(max_column+1)}{i}'] = f'=SUM({get_column_letter(min_column+1)}{i}:{get_column_letter(max_column)}{i})'
        sheet[f'{get_column_letter(max_column+1)}{i}'].style = 'Currency'
    sheet[f'{get_column_letter(max_column+1)}{min_row}'] = 'Total' 
    sheet[f'{get_column_letter(max_column+1)}{min_row}'].font = Font('Calibri',bold=True,size=11)
    wb.save(sample_file_location)

def create_chart():
    wb = load_workbook(sample_file_location)

    sheet = wb['Report']

    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row= wb.active.max_row

    data = Reference(sheet,min_col=min_column,max_col=max_column,
                            min_row = min_row,max_row=max_row)

    categories = Reference(sheet,min_col=min_column,max_col=min_column,
                                    min_row=min_row+1,max_row=max_row)
    
    barchart = BarChart()

    barchart.add_data(data,titles_from_data=True)
    barchart.set_categories(categories)
    barchart.title = 'Retail VS WholeSale'
    barchart.style = 40

    sheet.add_chart(barchart,"F1")
    wb.save(sample_file_location)
    
def addHeadlines():
    wb =load_workbook(sample_file_location)

    Title = 'Sales Report'
    sheet = wb['Report']
    sheet['A1'] = Title
    sheet['A1'].font = Font('Arial',bold=True,size=20)
    sheet['A2'] = month
    sheet['A2'].font = Font('Arial',bold=True,size=10)
    wb.save(sample_file_location)

create_sample_data(file_location)
create_chart()
print("Chart created!")
sumofEach()
print("Total for Each Category created!")
addHeadlines()
print("Headlines add!")
time.sleep(7)